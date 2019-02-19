from openpyxl import load_workbook

class VersionPrefix:
    def __init__(self, JiraHandler, url):
        self.jh = JiraHandler
        self.url = url

    # Add prefix based on the excel sheet and return a failed list
    def addPrefix():
        wb = load_workbook('out/version_name_changes_tom_luke_projects.xlsx')

        cmtd = []
        cs = []
        catm = []
        ssd = []
        failedList = []

        print("Start ARSATCMTD")
        ws = wb['ARSATCMTD']
        for row in ws.iter_rows(min_row=2):
            id = str(row[1].value)
            result = self.jh.getVersion(id)
            if result:
                update_result = self.jh.updateVersion(id,{'name':'CMTD ' + result.get('name')})
                if update_result == False:
                    failedList.append(id)
                print(update_result,id)
        print("Done")

        print("Start ARSATCS")
        ws = wb['ARSATCS']
        for row in ws.iter_rows(min_row=2):
            id = str(row[1].value)
            result = self.jh.getVersion(id)
            if result:
                update_result = self.jh.updateVersion(id,{'name':'CS ' + result.get('name')})
                if update_result == False:
                    failedList.append(id)
                print(update_result,id)
        print("Done")

        print("Start CATM")
        ws = wb['CATM']
        for row in ws.iter_rows(min_row=2):
            id = str(row[1].value)
            result = self.jh.getVersion(id)
            if result:
                update_result = self.jh.updateVersion(id,{'name':'CATM ' + result.get('name')})
                if update_result == False:
                    failedList.append(id)
                print(update_result,id)
        print("Done")

        print("Start SSD")
        ws = wb['SSD']
        for row in ws.iter_rows(min_row=2):
            id = str(row[1].value)
            result = self.jh.getVersion(id)
            if result:
                update_result = self.jh.updateVersion(id,{'name':'SSD ' + result.get('name')})
                if update_result == False:
                    failedList.append(id)
                print(update_result,id)
        print("Done")

        return failedList
